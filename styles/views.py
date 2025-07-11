from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
import openpyxl
import csv
from .models import Style, StyleFabricConsumption, StyleAccessoryConsumption
from masters.models import Fabric, Accessory

class StyleListView(LoginRequiredMixin, ListView):
    model = Style
    template_name = 'styles/style_list.html'
    context_object_name = 'styles'

class StyleCreateView(LoginRequiredMixin, CreateView):
    model = Style
    template_name = 'styles/style_form.html'
    fields = ['code', 'name', 'category', 'image']
    success_url = reverse_lazy('styles:style_list')

class StyleDetailView(LoginRequiredMixin, DetailView):
    model = Style
    template_name = 'styles/style_detail.html'
    context_object_name = 'style'

class StyleUpdateView(LoginRequiredMixin, UpdateView):
    model = Style
    template_name = 'styles/style_form.html'
    fields = ['code', 'name', 'category', 'image']
    success_url = reverse_lazy('styles:style_list')

class StyleDeleteView(LoginRequiredMixin, DeleteView):
    model = Style
    template_name = 'styles/style_confirm_delete.html'
    success_url = reverse_lazy('styles:style_list')


class StyleFabricConsumptionCreateView(LoginRequiredMixin, CreateView):
    model = StyleFabricConsumption
    template_name = 'styles/style_fabric_consumption_form.html'
    fields = ['fabric', 'quantity', 'unit']

    def form_valid(self, form):
        form.instance.style = Style.objects.get(pk=self.kwargs['style_pk'])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('styles:style_detail', kwargs={'pk': self.kwargs['style_pk']})

class StyleFabricConsumptionUpdateView(LoginRequiredMixin, UpdateView):
    model = StyleFabricConsumption
    template_name = 'styles/style_fabric_consumption_form.html'
    fields = ['fabric', 'quantity', 'unit']

    def get_success_url(self):
        return reverse_lazy('styles:style_detail', kwargs={'pk': self.kwargs['style_pk']})

class StyleFabricConsumptionDeleteView(LoginRequiredMixin, DeleteView):
    model = StyleFabricConsumption
    template_name = 'styles/style_fabric_consumption_confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('styles:style_detail', kwargs={'pk': self.kwargs['style_pk']})


class StyleAccessoryConsumptionCreateView(LoginRequiredMixin, CreateView):
    model = StyleAccessoryConsumption
    template_name = 'styles/style_accessory_consumption_form.html'
    fields = ['accessory', 'quantity', 'unit']

    def form_valid(self, form):
        form.instance.style = Style.objects.get(pk=self.kwargs['style_pk'])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('styles:style_detail', kwargs={'pk': self.kwargs['style_pk']})

class StyleAccessoryConsumptionUpdateView(LoginRequiredMixin, UpdateView):
    model = StyleAccessoryConsumption
    template_name = 'styles/style_accessory_consumption_form.html'
    fields = ['accessory', 'quantity', 'unit']

    def get_success_url(self):
        return reverse_lazy('styles:style_detail', kwargs={'pk': self.kwargs['style_pk']})

class StyleAccessoryConsumptionDeleteView(LoginRequiredMixin, DeleteView):
    model = StyleAccessoryConsumption
    template_name = 'styles/style_accessory_consumption_confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('styles:style_detail', kwargs={'pk': self.kwargs['style_pk']})

class ExportConsumptionView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="consumption_report.xlsx"'

        workbook = openpyxl.Workbook()

        # Fabric Consumption Sheet
        fabric_sheet = workbook.active
        fabric_sheet.title = "Fabric Consumption"
        fabric_sheet.append(['Style Code', 'Style Name', 'Fabric Name', 'Quantity', 'Unit'])

        fabric_consumptions = StyleFabricConsumption.objects.all().select_related('style', 'fabric')
        for fc in fabric_consumptions:
            fabric_sheet.append([
                fc.style.code,
                fc.style.name,
                fc.fabric.name,
                fc.quantity,
                fc.unit
            ])

        # Accessory Consumption Sheet
        accessory_sheet = workbook.create_sheet("Accessory Consumption")
        accessory_sheet.append(['Style Code', 'Style Name', 'Accessory Name', 'Quantity', 'Unit'])

        accessory_consumptions = StyleAccessoryConsumption.objects.all().select_related('style', 'accessory')
        for ac in accessory_consumptions:
            accessory_sheet.append([
                ac.style.code,
                ac.style.name,
                ac.accessory.name,
                ac.quantity,
                ac.unit
            ])
        
        workbook.save(response)
        return response

class ImportConsumptionView(LoginRequiredMixin, View):
    def get(self, request):
        form = ConsumptionUploadForm()
        return render(request, 'styles/import_consumption.html', {'form': form})

    def post(self, request):
        form = ConsumptionUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            
            # Determine file type and read accordingly
            if file.name.endswith('.csv'):
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
            elif file.name.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                # Assuming first row is header
                headers = [cell.value for cell in sheet[1]]
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data.append(dict(zip(headers, row)))
                reader = data # Treat as a list of dictionaries
            else:
                messages.error(request, "Unsupported file type. Please upload a CSV or XLSX file.")
                return render(request, 'styles/import_consumption.html', {'form': form})

            # Assuming CSV/Excel format: Style Code, Item Type (Fabric/Accessory), Item Name, Quantity, Unit
            for row in reader:
                style_code = row.get('Style Code')
                item_type = row.get('Item Type')
                item_name = row.get('Item Name')
                quantity = row.get('Quantity')
                unit = row.get('Unit')

                try:
                    style = Style.objects.get(code=style_code)
                    quantity = float(quantity)

                    if item_type == 'Fabric':
                        fabric = Fabric.objects.get(name=item_name)
                        StyleFabricConsumption.objects.update_or_create(
                            style=style,
                            fabric=fabric,
                            defaults={'quantity': quantity, 'unit': unit}
                        )
                    elif item_type == 'Accessory':
                        accessory = Accessory.objects.get(name=item_name)
                        StyleAccessoryConsumption.objects.update_or_create(
                            style=style,
                            accessory=accessory,
                            defaults={'quantity': quantity, 'unit': unit}
                        )
                    else:
                        messages.warning(request, f"Unknown item type '{item_type}' for row: {row}. Skipping.")
                except Style.DoesNotExist:
                    messages.error(request, f"Style with code '{style_code}' not found for row: {row}. Skipping.")
                except Fabric.DoesNotExist:
                    messages.error(request, f"Fabric with name '{item_name}' not found for row: {row}. Skipping.")
                except Accessory.DoesNotExist:
                    messages.error(request, f"Accessory with name '{item_name}' not found for row: {row}. Skipping.")
                except ValueError:
                    messages.error(request, f"Invalid quantity '{quantity}' for row: {row}. Skipping.")
                except Exception as e:
                    messages.error(request, f"Error processing row: {row} - {e}. Skipping.")

            messages.success(request, "Consumption data import process completed. Check for any errors above.")
            return redirect('styles:style_list')
        return render(request, 'styles/import_consumption.html', {'form': form})
